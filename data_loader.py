"""
AgentBench dataset loader with context expansion for 32k compression testing.

Strategy (Research Background Document Injection):
- Insert a medium-sized "research background document" as chunked
    user/assistant message pairs between the system prompt and the first real
    conversation turn.
- This document becomes part of B1 (compressible history), NOT A (system).
- The compressor summarizes the reference material; B2/B1 is diagnostic only.
- Content is generic "lateral thinking puzzle solving techniques" — does NOT
  contain puzzle-specific answers, so benchmark accuracy is preserved.
- Realistic: agents often have large reference documents / briefing materials.

ABC structure:
  Before compression:
    [A: system prompt]
    + [B1: research chunks + accumulated dialogue]
    + [C1: KEEP_RECENT_TOKENS_BUDGET ~2-3k]
    After compression:
    [A: system prompt]
    + [B2: summary of research + early dialogue]
    + [C1: recent turns ~2-3k]
"""

import json
import os
import random
import sys
from copy import deepcopy
from pathlib import Path
from typing import List, Optional

BENCH_ROOT = Path(__file__).parent.resolve()
if str(BENCH_ROOT) not in sys.path:
    sys.path.insert(0, str(BENCH_ROOT))
sys.path.append(str(Path(__file__).parent.parent / "bfcl_compression_bench"))
from compressor import estimate_tokens, estimate_message_tokens

import config as CFG

# ============================================================
# Research background document — "Lateral Thinking Puzzle Guide"
# ============================================================
#
# This document is a generic guide on how to solve lateral thinking puzzles.
# It is deliberately long (~20-25k tokens) with varied content so that the
# compressor can produce a meaningful structured summary for diagnostics.
#
# Each section covers different aspects: history, common patterns,
# reasoning techniques, example analysis, recommended question types, etc.
# The content is VARIED (not repetitive) to achieve REALISTIC compression
# ratios — repetitive text would give artificially high compression.
#

RESEARCH_DOC_SECTIONS = [
    # ---- Section 1: Introduction & History ----
    (
        "## 1. History and Origins of Lateral Thinking Puzzles",
        """
Lateral thinking puzzles, also known as "situation puzzles" or "mystery puzzles,"
originated from the concept of lateral thinking developed by Edward de Bono in
1967. De Bono argued that traditional vertical thinking (step-by-step logical
deduction) is insufficient for solving certain types of problems that require
creative insight and reframing of assumptions.

The puzzle format became popular in the 1980s through Paul Sloane's collections,
which introduced the classic "Yes/No/Irrelevant" host-response format. In Japan,
these puzzles are known as "Umi no Kame no Soup" (海龟汤), named after a famous
puzzle about a man who orders turtle soup at a restaurant and then commits suicide
after tasting it.

The genre has since expanded globally, with dedicated puzzle communities,
competition events, and thousands of documented puzzles across difficulty levels.
Modern implementations include AI-hosted puzzles where language models serve as
the game host, providing structured responses to player inquiries.

Key characteristics that define the genre:
- A brief, mysterious narrative setup (often 50-200 words)
- A hidden explanation that resolves the apparent contradiction
- Players ask yes/no questions to the host
- The host can only respond with "Yes," "No," or "Irrelevant"
- Solutions often involve unexpected interpretations of common assumptions
""",
    ),
    # ---- Section 2: Common Patterns ----
    (
        "## 2. Common Puzzle Patterns and Solution Archetypes",
        """
Understanding common patterns is essential for efficient puzzle solving. Over
decades of puzzle creation, several recurring archetypes have been identified:

### 2.1 Misinterpretation of Identity
The most common pattern involves mistaken identity: a character is not who they
appear to be. Classic examples include:
- A "waiter" who is actually a patient in a psychiatric hospital
- A "businessman" who is actually a prisoner on work release
- A "tourist" who is actually a spy surveilling a target location
- A "student" who is actually an undercover police officer

### 2.2 Unexpected Environmental Context
The setting contains hidden information that changes the interpretation:
- A room described as "dark" because the protagonist is blind
- A building "in the middle of nowhere" that is actually on a military base
- A "quiet neighborhood" that is actually a cemetery
- An "empty room" that is actually a prison cell during a riot

### 2.3 Hidden Relationships
Characters have connections that are not immediately obvious:
- The victim and perpetrator are the same person (suicide disguised as murder)
- Two seemingly unrelated people are actually siblings separated at birth
- A person's "employer" is actually their blackmailer
- The "stranger" who helps is actually a relative in disguise

### 2.4 Temporal or Causal Reversal
The sequence of events is different from what is assumed:
- The "cause" happens AFTER the "effect" in the narrative
- An action described as malicious was actually accidental and regretted
- The "crime scene" was staged after the fact to mislead investigators
- A character described as "fleeing" was actually running TOWARD help

### 2.5 Semantic Ambiguity
Words in the puzzle description have multiple meanings:
- "Bank" could mean financial institution or river bank
- "Ring" could mean jewelry, circle, or boxing ring
- "Light" could mean illumination or not heavy
- "Right" could mean direction or correctness

### 2.6 Scale Mismatch
The scale of objects or events is misunderstood:
- A "ship in a bottle" that is actually a message in a bottle
- A "mountain" that is actually a molehill viewed from a miniature perspective
- A "crowd of thousands" that is actually an ant colony
- "Swimming across the ocean" that is actually swimming across a swimming pool
""",
    ),
    # ---- Section 3: Reasoning Methodology ----
    (
        "## 3. Systematic Reasoning Methodology for Lateral Thinking Puzzles",
        """
### 3.1 The Four-Phase Approach

**Phase 1: Surface Analysis (Turns 1-5)**
Begin by cataloging every explicit element in the story:
- All named and unnamed characters
- All described locations and objects
- All stated actions and events
- All temporal references (time of day, season, duration)
- All emotional descriptors (surprised, sad, afraid, angry)

For each element, list at least 2-3 possible interpretations. For example,
if the story mentions "a man walks into a bar," consider: (a) a drinking
establishment, (b) a metal rod, (c) a legal examination room, (d) a chocolate
bar (e) a sandbar on a beach.

**Phase 2: Anomaly Identification (Turns 5-10)**
Identify what makes this story puzzling or unusual:
- What is the central contradiction or surprise?
- What would be the "normal" version of this story?
- What assumptions does the "normal" interpretation rely on?
- Which of these assumptions can be challenged?

**Phase 3: Hypothesis Generation (Turns 10-18)**
For each anomaly, generate multiple explanatory hypotheses:
- What identity assumptions could be wrong?
- What environmental assumptions could be wrong?
- What causal assumptions could be wrong?
- What semantic assumptions could be wrong?

Rank hypotheses by consistency with known facts. Eliminate those directly
contradicted by host responses. Refine remaining hypotheses with follow-up
questions.

**Phase 4: Verification and Synthesis (Turns 18-25)**
Once a leading hypothesis emerges:
- Design targeted yes/no questions that differentiate between remaining hypotheses
- Ask about specific details that would confirm or refute each possibility
- Build a complete narrative that accounts for ALL stated facts
- Verify the narrative against each element of the original story

### 3.2 Question Design Principles

**Good questions are:**
1. **Specific**: "Was the man in the bar there to drink alcohol?" > "Was he there for a reason?"
2. **Falsifiable**: Each question should have the potential to eliminate a hypothesis
3. **Independent**: Don't ask compound questions that require multiple answers
4. **Progressive**: Questions should build on previously confirmed facts
5. **Gap-filling**: Target the most important unknown in your current theory

**Bad questions to avoid:**
1. "Did something unusual happen?" — Too vague, host will clarify with follow-up
2. "Can you tell me the answer?" — Against game rules, host will refuse
3. "Is my theory correct?" — Premature; verify components individually
4. "Was it X or Y?" — This is two questions; ask one at a time
5. "Is there something I'm missing?" — Always true, provides no information

### 3.3 Information Theory Perspective

Each yes/no answer provides exactly 1 bit of information if the question
is well-designed (50/50 prior probability). Poor questions provide fractional
bits. The goal is to maximize information gained per turn.

For a puzzle with 5 critical answer keys, assuming each key has ~8 possible
states, the total uncertainty is approximately 5 * log2(8) = 15 bits. With
efficient questioning, this can be resolved in 15-20 turns.
""",
    ),
    # ---- Section 4: Example Puzzles with Analysis ----
    (
        "## 4. Annotated Example Puzzles",
        """
### Example 1: "The Man in the Bar"
**Story**: A man walks into a bar and asks for a glass of water. The bartender
pulls out a gun and points it at him. The man says "Thank you" and leaves.
Why?

**Analysis**:
The anomaly here is the man thanking the bartender for threatening him. Surface
analysis reveals: (a) a bar, (b) a glass of water, (c) a gun, (d) gratitude.
Normal interpretation: asking for water at a bar is unusual; being threatened
with a gun should provoke fear, not gratitude.

**Solution path**: The key is recognizing that "water" solves a specific problem.
The man had hiccups and wanted water to cure them. The bartender's gun startled
him, which also cures hiccups. The man thanked the bartender for curing his
hiccups (by startling him), not for the water (which he no longer needed).

**Reasoning pattern**: This puzzle uses unexpected causal relationship — the
fear response from the gun cures the hiccups, so the apparent threat was actually
helpful. This illustrates the "hidden benefit" subtype of temporal reversal.

### Example 2: "The Albatross"
**Story**: A man orders albatross soup at a restaurant. After tasting one spoonful,
he leaves the restaurant and commits suicide. Why?

**Analysis**:
Surface elements: man, restaurant, albatross soup, tasting, suicide.
Key anomaly: tasting soup leads to suicide — an extreme reaction to food.

**Solution path**: The man was a shipwreck survivor. He and his companions were
stranded on an island where his friend died. The other survivors told him they
were eating albatross to survive. Now, tasting real albatross, he realizes the
"albatross" he ate before was actually his deceased friend. The horror of this
realization drives him to suicide.

**Reasoning pattern**: Semantic ambiguity ("albatross" = real bird vs. euphemism
for human flesh) combined with a delayed realization. This is the classic
"semantic ambiguity" archetype.

### Example 3: "The Elevator"
**Story**: A man lives on the 10th floor. Every morning he takes the elevator
down to the lobby and goes to work. When he returns, if it's raining he takes
the elevator to the 10th floor. If it's not raining, he takes the elevator to
the 5th floor and walks up the stairs to the 10th floor. Why?

**Analysis**:
Surface elements: apartment building, elevator, rain, 5th floor, stairs.
Key anomaly: behavior changes based on weather — affecting building navigation.

**Solution path**: The man is a dwarf (or very short person). He cannot reach the
button for the 10th floor. When it's raining, he uses his umbrella to press the
10 button. When it's not raining, he can only reach the 5 button, so he gets off
there and walks up the rest of the way.

**Reasoning pattern**: Hidden physical characteristic (short stature) combined
with a tool that only becomes available in specific conditions (umbrella when
raining). This illustrates the "hidden attribute + conditional tool" pattern.

### Example 4: "The Cabin in the Woods"
**Story**: A man is found dead in a cabin in the middle of a forest. There are
no signs of a struggle. The cabin contains only a table and a chair. On the
table is an unopened envelope. How did he die?

**Analysis**:
Surface elements: dead man, isolated cabin, no struggle, table, chair, unopened
envelope. Key anomaly: no obvious cause of death, nothing apparently dangerous.

**Solution path**: The cabin is the cabin of an airplane that crashed in the
forest. The man died in the crash. The "cabin" is an airplane cabin, not a
wooden cabin. The "envelope" is the flight envelope (operating parameters of
the aircraft), not a paper envelope.

**Reasoning pattern**: Extreme semantic ambiguity where both "cabin" and
"envelope" have unexpected meanings in an aviation context. This illustrates
how multiple semantic ambiguities can combine in a single puzzle.
""",
    ),
    # ---- Section 5: Statistical Analysis of Puzzle Solving ----
    (
        "## 5. Statistical Patterns in Effective Puzzle Solving",
        """
### 5.1 Question Type Effectiveness

Analysis of 500+ solved puzzles reveals the following effectiveness hierarchy
for question types (measured by information gain per question):

| Question Type                    | Avg Info Gain | Success Rate |
|----------------------------------|---------------|--------------|
| Identity of characters           | 0.87 bits     | 72%          |
| Relationship between characters  | 0.82 bits     | 68%          |
| Location/setting clarification   | 0.79 bits     | 65%          |
| Temporal sequence verification   | 0.76 bits     | 63%          |
| Motivation/intent of actions     | 0.73 bits     | 60%          |
| Physical properties of objects   | 0.71 bits     | 58%          |
| Semantic meaning of words        | 0.68 bits     | 55%          |
| Background/history of characters | 0.64 bits     | 52%          |
| Abstract/general clarification   | 0.45 bits     | 35%          |
| Direct guess of answer           | 0.30 bits     | 20%          |

### 5.2 Optimal Turn Allocation

For a typical 25-turn puzzle, the optimal allocation across phases is:
- Phase 1 (Surface Analysis): 4-6 turns (20-25% of total)
- Phase 2 (Anomaly Identification): 5-7 turns (20-25% of total)
- Phase 3 (Hypothesis Generation): 8-10 turns (35-40% of total)
- Phase 4 (Verification): 4-6 turns (15-20% of total)

Deviating from this allocation typically reduces solve rates:
- Too few surface analysis turns → missed critical details → wrong hypotheses
- Too many hypothesis turns without verification → inefficient use of turns
- Jumping to verification too early → premature commitment to wrong theory

### 5.3 Common Failure Modes

1. **Premature Convergence (35% of failures)**: Fixating on a hypothesis too
   early and asking confirmatory rather than exploratory questions.

2. **Scope Neglect (25% of failures)**: Focusing exclusively on one aspect of
   the puzzle while ignoring other potentially relevant elements.

3. **Confirmation Bias Loop (20% of failures)**: Interpreting "Irrelevant"
   responses as confirmatory and doubling down on wrong hypotheses.

4. **Key Detail Oversight (15% of failures)**: Missing a critical word or
   phrase in the original story that would have redirected the inquiry.

5. **Question Design Failure (5% of failures)**: Asking questions that are
   too vague or compound, receiving ambiguous feedback, and proceeding with
   incorrect understanding.
""",
    ),
    # ---- Section 6: Advanced Techniques ----
    (
        "## 6. Advanced Lateral Thinking Techniques",
        """
### 6.1 The Assumption Inversion Method

For each assumption you hold about the puzzle, explicitly write it down, then
systematically invert it:

Original assumption → Inverted version → Test question

Example:
- Assumption: "The man is alive at the beginning" → "The man is already dead" →
  "Was the man already deceased before the events described?"
- Assumption: "The location is on Earth" → "The location is not on Earth" →
  "Does this take place in a non-terrestrial setting?"
- Assumption: "The events are in chronological order" → "Events are told out of
  order" → "Did the first event described actually happen last?"

### 6.2 The Five Whys Technique

Adapted from root cause analysis: for each puzzling element, ask "Why?" five
times, with each answer becoming the premise for the next question.

Example for "The man thanks the bartender":
1. Why does the man thank the bartender? → The gun solved a problem
2. Why does a gun solve a problem? → It causes a physiological response
3. Why is the physiological response desired? → It cures a condition
4. Why did the man have this condition? → [Follow-up question for host]
5. Why did he go to a bar for this? → Because bars are accessible public places

### 6.3 Constraint-Based Reasoning

Define the solution space by negative constraints:

For each host response of "No" or "Irrelevant," record what is EXCLUDED:
- Host said "No" to supernatural cause → All supernatural explanations eliminated
- Host said "Irrelevant" to weather conditions → Weather is not a factor
- Host said "No" to multiple perpetrators → Only one person involved

The intersection of all negative constraints defines the remaining solution space.
As constraints accumulate, the space shrinks until only the correct solution remains.

### 6.4 Analogical Transfer

Map the current puzzle to structurally similar solved puzzles:

1. Identify the deep structure: what TYPE of anomaly is present?
2. Search memory for puzzles with similar deep structure
3. Transfer the solution framework, not the specific answer
4. Adapt the framework to the specific details of the current puzzle

For example, if the anomaly involves "unexpected gratitude," recall puzzles
where apparent harm was actually beneficial (The Man in the Bar). The deep
structure is "hidden benefit from apparent threat" — this framework can be
tested against the current puzzle without assuming the same specific mechanism.

### 6.5 The Devil's Advocate Protocol

Periodically (every 5-7 turns), consciously argue AGAINST your current leading
hypothesis:
1. "If my theory is wrong, what would the real explanation look like?"
2. "What evidence would definitively disprove my current theory?"
3. "What alternative theories have I prematurely dismissed?"
4. "Am I interpreting 'Irrelevant' as 'consistent with my theory'?"
5. "What question should I ask that might break my theory?"

This prevents premature convergence and maintains exploratory diversity.
""",
    ),
    # ---- Section 7: Question Taxonomy ----
    (
        "## 7. Comprehensive Question Taxonomy for Lateral Thinking",
        """
### 7.1 Identity Questions
- "Is [character] actually a [profession/role]?"
- "Does [character] know [other character]?"
- "Is [character] who they appear to be?"
- "Is [character] lying about their identity?"
- "Is [character] related to [other character]?"

### 7.2 Location & Setting Questions
- "Does this take place in a [specific location type]?"
- "Is the setting important to understanding the anomaly?"
- "Is there something unusual about the physical environment?"
- "Does the time period matter?"
- "Is the location different from what it appears to be?"

### 7.3 Causal Questions
- "Did [event A] cause [event B] or vice versa?"
- "Is the apparent cause actually an effect?"
- "Is there a hidden third factor causing both [A] and [B]?"
- "Would the same outcome have occurred without [factor X]?"
- "Is the causal relationship coincidental rather than meaningful?"

### 7.4 Semantic Questions
- "Does [word] have a non-standard meaning here?"
- "Is [word] being used as a technical term from a specific domain?"
- "Could [phrase] be interpreted literally rather than figuratively?"
- "Does [word] refer to something other than its most common meaning?"
- "Is there wordplay or punning involved?"

### 7.5 Motivational Questions
- "Did [character] intend the outcome that occurred?"
- "Was [character]'s action voluntary or coerced?"
- "Did [character] have knowledge that would change their behavior?"
- "Is [character]'s stated reason for their action truthful?"
- "Was the action motivated by fear, greed, love, duty, or something else?"

### 7.6 Counterfactual Questions
- "If [character] had known [fact], would things have been different?"
- "If [event] had not happened, would the outcome be the same?"
- "Is there an alternative sequence of events that explains everything?"
- "Would a reasonable person in [character]'s position act the same way?"
""",
    ),
    # ---- Section 8: Research Summary ----
    (
        "## 8. Research Summary and Best Practices",
        """
### 8.1 Key Takeaways

1. **Systematic approach beats intuition**: Solvers who follow a structured
   methodology (like the four-phase approach) consistently outperform those
   who rely on intuition alone, by approximately 35% in controlled studies.

2. **Question efficiency is the primary differentiator**: The top quartile of
   solvers achieves approximately 0.8 bits of information per question, while
   the bottom quartile achieves only 0.3 bits. This 2.7x efficiency gap
   translates directly to solving speed and success rates.

3. **Assumption documentation prevents oversight**: Solvers who explicitly
   write down their assumptions and systematically test them miss 60% fewer
   critical details than those who keep assumptions implicit.

4. **Early divergence, late convergence**: The most successful solving pattern
   involves broad exploration in the first 60% of turns, followed by rapid
   convergence in the final 40%. Premature convergence is the single largest
   source of failure.

5. **"Irrelevant" responses are data**: An "Irrelevant" response is not a
   waste of a turn — it provides valuable negative information that constrains
   the remaining solution space. Track irrelevance patterns to identify dead ends.

### 8.2 Recommended Reading

For further study, the following resources are recommended:
- De Bono, E. "Lateral Thinking: Creativity Step by Step" (1970)
- Sloane, P. "Lateral Thinking Puzzlers" (1992)
- Sloane, P. & MacHale, D. "Great Lateral Thinking Puzzles" (1994)
- Bodycombe, D. "The Lateral Thinking Quiz Book" (2008)
- The Situation Puzzle Community Wiki (online resource)

### 8.3 Practice Recommendations

To improve puzzle-solving skills:
1. Solve 3-5 puzzles per week, documenting your question strategy
2. Review solved puzzles to identify missed optimal question paths
3. Practice with puzzles of increasing difficulty (1-5 star ratings)
4. Join puzzle-solving communities to learn diverse approaches
5. Record your personal "patterns recognized" library for future reference
""",
    ),
]

# ============================================================
# Core expansion logic
# ============================================================


def build_research_document(sections: Optional[list] = None,
                            standalone: bool = True) -> str:
    """Build the full research document from predefined sections.

    Args:
        sections: Override sections to include (None = all).
        standalone: If True, add wrapper context explaining the document.

    Returns:
        Full research document as a single string.
    """
    if sections is None:
        sections = RESEARCH_DOC_SECTIONS

    parts = []
    if standalone:
        parts.append(
            "=== RESEARCH REFERENCE: Lateral Thinking Puzzle Solving Guide ===\n"
            "The following is a comprehensive reference document on lateral\n"
            "thinking puzzle methodology, compiled from academic research and\n"
            "expert practitioner experience. Study this material carefully\n"
            "before beginning the puzzle.\n"
        )

    for title, content in sections:
        parts.append(f"{title}\n{content}")

    return "\n".join(parts)


def create_research_message_pair(
    doc_text: str,
    target_tokens: int = None,
    tokenizer=None,
) -> tuple[dict, dict]:
    """Create a user/assistant message pair containing the research document.

    The user message contains the full research document. The assistant message
    is a brief acknowledgment. This pair becomes part of B1 and is compressible.

    If target_tokens is specified and tokenizer is provided, the function will
    ensure the document is long enough to meet the target.

    Returns (user_msg, assistant_msg).
    """
    if target_tokens and tokenizer:
        # Expand if needed
        current_tokens = len(tokenizer.tokenize(doc_text))
        if current_tokens < target_tokens:
            # Replicate sections with different numbering to reach target
            expanded = [doc_text]
            section_idx = len(RESEARCH_DOC_SECTIONS) + 1
            seed = 42 + section_idx
            rng = random.Random(seed)

            while current_tokens < target_tokens:
                # Add a variant section
                more = build_expansion_section(section_idx, rng)
                expanded.append(more)
                current_tokens = len(tokenizer.tokenize("\n".join(expanded)))
                section_idx += 1

            doc_text = "\n".join(expanded)

    user_msg = {
        "role": "user",
        "content": (
            "[This section contains reference material for the task. "
            "This is background research documentation, NOT part of the "
            "current puzzle conversation. Use this information to inform "
            "your puzzle-solving strategy.]\n\n" + doc_text
        ),
    }

    assistant_msg = {
        "role": "assistant",
        "content": (
            "I have carefully reviewed the reference material on lateral "
            "thinking puzzle solving techniques. I understand the four-phase "
            "approach, the common archetype patterns, and the recommended "
            "question design principles. I am ready to apply these strategies "
            "to the puzzle."
        ),
    }

    return user_msg, assistant_msg


def split_research_message_pair(
    user_msg: dict,
    assistant_msg: dict,
    tokenizer,
    max_chunk_tokens: int = 900,
) -> list[dict]:
    """Split a large research document into compressible dialogue turns."""
    content = str(user_msg.get("content", ""))
    paragraphs = content.split("\n\n")
    messages = []
    chunk_parts = []
    chunk_idx = 1

    def token_count(text: str) -> int:
        return len(tokenizer.tokenize(text)) if tokenizer is not None else max(1, len(text) // 4)

    def flush_chunk():
        nonlocal chunk_parts, chunk_idx
        if not chunk_parts:
            return
        chunk_text = "\n\n".join(chunk_parts)
        messages.append({
            "role": "user",
            "content": f"[Reference material chunk {chunk_idx}]\n\n{chunk_text}",
        })
        messages.append({
            "role": "assistant",
            "content": (
                f"Acknowledged reference material chunk {chunk_idx}. "
                "I will retain it as background guidance for the puzzle."
            ),
        })
        chunk_parts = []
        chunk_idx += 1

    for paragraph in paragraphs:
        candidate_parts = chunk_parts + [paragraph]
        candidate = "\n\n".join(candidate_parts)
        if chunk_parts and token_count(candidate) > max_chunk_tokens:
            flush_chunk()
        chunk_parts.append(paragraph)

    flush_chunk()

    if not messages:
        return [user_msg, assistant_msg]

    messages.append(assistant_msg)
    return messages


def build_expansion_section(idx: int, rng: random.Random) -> str:
    """Build a single appendix section for document expansion.

    Generates varied content so the document remains compressible at a
    realistic ratio (~1/4 to 1/5) rather than being purely repetitive.
    """
    templates = [
        (
            f"## Appendix {idx}: Additional Case Study Analysis",
            f"This appendix presents supplementary case studies analyzed "
            f"using the methodological framework described above.\n\n"
            f"### Case {idx}a: Pattern Recognition Study\n"
            f"A corpus of {rng.randint(200,500)} puzzles was analyzed for "
            f"recurring pattern frequencies across different difficulty "
            f"levels. The distribution revealed that semantic ambiguity "
            f"patterns account for {rng.randint(25,40)}% of puzzles rated "
            f"3-star and above, while identity misinterpretation dominates "
            f"among 1-2 star puzzles at {rng.randint(35,50)}%.\n\n"
            f"### Case {idx}b: Cross-Cultural Comparison\n"
            f"Comparing puzzle-solving approaches across {rng.randint(3,8)} "
            f"cultural contexts revealed significant variations in question "
            f"formulation strategies. Participants from cultures with higher "
            f"context communication styles tended to ask {rng.randint(15,30)}% "
            f"more relationship-focused questions than those from low-context "
            f"cultures.\n\n"
            f"### Case {idx}c: Expert vs. Novice Comparison\n"
            f"Analysis of {rng.randint(50,150)} expert solvers (500+ puzzles "
            f"solved) versus {rng.randint(150,400)} novices (fewer than 20 "
            f"puzzles solved) showed that experts spend {rng.randint(40,60)}% "
            f"more time in the surface analysis phase and ask questions with "
            f"{rng.randint(30,50)}% higher information density."
        ),
        (
            f"## Appendix {idx}: Supplementary Reasoning Frameworks",
            f"### Framework {idx}a: Multi-Hypothesis Tracking\n"
            f"Maintain {rng.randint(3,7)} active hypotheses simultaneously. "
            f"After each host response, update the probability of each "
            f"hypothesis using Bayesian reasoning. Eliminate hypotheses "
            f"when their probability drops below {rng.randint(5,15)}%.\n\n"
            f"### Framework {idx}b: Decision Tree Pruning\n"
            f"Construct a decision tree where each node is a binary question "
            f"and each branch represents a Yes/No response. Use alpha-beta "
            f"pruning to eliminate branches that cannot lead to the correct "
            f"solution within the remaining turn budget.\n\n"
            f"### Framework {idx}c: Information Gain Maximization\n"
            f"For each candidate question, compute the expected information "
            f"gain as: IG(Q) = -∑ P(r|Q) * log2(P(r|Q)) where r ranges over "
            f"{{Yes, No, Irrelevant}}. Select questions that maximize expected "
            f"information gain while minimizing the risk of 'Irrelevant' "
            f"responses."
        ),
        (
            f"## Appendix {idx}: Puzzle Database Statistics",
            f"### Database Overview\n"
            f"This appendix summarizes statistical analysis of a curated "
            f"database of {rng.randint(5000,15000)} lateral thinking puzzles "
            f"from {rng.randint(8,20)} different sources.\n\n"
            f"### Difficulty Distribution\n"
            f"- 1-star (easiest): {rng.randint(15,25)}% of database\n"
            f"- 2-star: {rng.randint(25,35)}% of database\n"
            f"- 3-star: {rng.randint(20,30)}% of database\n"
            f"- 4-star: {rng.randint(10,20)}% of database\n"
            f"- 5-star (hardest): {rng.randint(5,10)}% of database\n\n"
            f"### Average Solving Metrics\n"
            f"- Mean turns to solve 3-star puzzles: {rng.randint(15,22)}\n"
            f"- Median questions asked: {rng.randint(18,28)}\n"
            f"- Information efficiency: {rng.uniform(0.55, 0.85):.2f} bits/question\n"
            f"- First-attempt solve rate: {rng.randint(55,75)}%\n"
            f"- Host 'Irrelevant' rate: {rng.randint(15,30)}%"
        ),
    ]

    idx_mod = idx % len(templates)
    title, body = templates[idx_mod]
    # Modify the body slightly with idx to ensure uniqueness
    return f"{title}\n{body}"


# ============================================================
# Message list construction
# ============================================================


def build_initial_messages(
    system_prompt: str,
    host_system_prompt: str = None,
    target_initial_tokens: int = None,
    tokenizer=None,
    include_research: bool = True,
) -> list[dict]:
    """Build initial message list with optional research document injection.

    Structure:
      [system: system_prompt]
      [user: research_document]        ← only if include_research=True
      [assistant: acknowledgment]      ← only if include_research=True

    The research pair becomes B1 material (compressible), NOT system (A).

    Args:
        system_prompt: The LTP agent system prompt.
        host_system_prompt: (Unused in messages, kept for compatibility).
        target_initial_tokens: Target token count for the full initial context
                               (A + research pair). Default from CFG.
        tokenizer: HuggingFace tokenizer for token estimation.
        include_research: Whether to include the research document pair.

    Returns:
        List of message dicts (OpenAI format).
    """
    if target_initial_tokens is None:
        target_initial_tokens = CFG.TARGET_INITIAL_TOKENS

    messages = [{"role": "system", "content": system_prompt}]

    if include_research and target_initial_tokens > 0 and tokenizer is not None:
        doc = build_research_document()
        research_user, research_assistant = create_research_message_pair(
            doc,
            target_tokens=target_initial_tokens,
            tokenizer=tokenizer,
        )
        messages.extend(split_research_message_pair(
            research_user,
            research_assistant,
            tokenizer=tokenizer,
        ))

    return messages


# ============================================================
# LTP puzzle loading
# ============================================================


def load_ltp_puzzles(
    filepath=None,
    limit: Optional[int] = None,
    target_initial_tokens=None,
    tokenizer=None,
    include_research=True,
) -> list[dict]:
    """Load LTP puzzles with optional research-document-based expansion.

    Each puzzle gets:
      - system_prompt: expanded agent system (A, unchanged by compressor)
      - host_system_prompt: host-facing system
      - initial_messages: [system] + [research_user, research_assistant]
    """
    if filepath is None:
        filepath = CFG.TASK_CATEGORIES["ltp"]["data_file"]
    if target_initial_tokens is None:
        target_initial_tokens = CFG.TARGET_INITIAL_TOKENS

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(str(filepath))
    ws = wb.active
    puzzles = []

    from run_llama_ltp_prompts import (
        AGENT_SYSTEM_PROMPT as LTP_AGENT_SYS,
        HOST_SYSTEM_PROMPT as LTP_HOST_SYS,
    )

    max_row = ws.max_row + 1
    if limit is not None:
        max_row = min(max_row, 2 + max(limit, 0))

    for i in range(2, max_row):
        story = ws.cell(i, 1).value
        answer = ws.cell(i, 2).value
        story_keys = ws.cell(i, 3).value
        answer_keys = ws.cell(i, 4).value
        if not story or not answer:
            continue

        puzzle = {
            "id": i - 2,
            "story": str(story).strip(),
            "answer": str(answer).strip(),
            "story_keys": str(story_keys or "").strip(),
            "answer_keys": str(answer_keys or "").strip(),
        }

        agent_sys_content = LTP_AGENT_SYS.format(
            story=puzzle["story"],
            max_turns=CFG.TASK_CATEGORIES["ltp"]["max_turns"],
        )
        host_sys_content = LTP_HOST_SYS.format(
            story=puzzle["story"],
            answer=puzzle["answer"],
            answer_keys=puzzle["answer_keys"],
        )

        puzzle["system_prompt"] = agent_sys_content
        puzzle["host_system_prompt"] = host_sys_content

        # Build initial messages with research doc injection
        puzzle["messages"] = build_initial_messages(
            agent_sys_content,
            host_sys_content,
            target_initial_tokens=target_initial_tokens,
            tokenizer=tokenizer,
            include_research=include_research,
        )

        puzzles.append(puzzle)

    return puzzles


# ============================================================
# Utilities
# ============================================================


def estimate_initial_token_distribution(
    puzzle: dict,
    tokenizer,
) -> dict:
    """Estimate the A / B1 / C1 token distribution for a puzzle's initial state.

    Returns dict with A_tokens, B1_tokens, C1_tokens — useful for verifying
    that the research doc is correctly positioned in B1 (compressible) not A.
    """
    msgs = puzzle.get("messages", [])
    if not msgs:
        return {"A_tokens": 0, "B1_tokens": 0, "C1_tokens": 0}

    # A = system message
    a_tokens = len(tokenizer.tokenize(msgs[0].get("content", "")))

    # B1 = everything after system
    b1_tokens = 0
    for m in msgs[1:]:
        b1_tokens += len(tokenizer.tokenize(m.get("content", "")))

    return {
        "A_tokens": a_tokens,
        "B1_tokens": b1_tokens,
        "total_initial": a_tokens + b1_tokens,
        "A_ratio": round(a_tokens / max(a_tokens + b1_tokens, 1) * 100, 1),
        "B1_ratio": round(b1_tokens / max(a_tokens + b1_tokens, 1) * 100, 1),
    }
